from flask import (
    Flask,
    render_template,
    request,
    flash,
    redirect,
    send_file,
    url_for,
    session,
)
from werkzeug.utils import secure_filename
import re
import os
import genetic_algorithm
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "llave_secreta"

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER


def is_number(value):
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def process_file(filepath):
    workbook = load_workbook(filename=filepath, read_only=True)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        row_tuple = []
        for cell in row:
            if cell is not None:
                if str(cell) == "2000c":
                    cell = 2000
                elif is_number(cell):
                    cell = float(cell) if "." in str(cell) else int(cell)
                else:
                    cell = str(cell)
                row_tuple.append(cell)
        if any(is_number(cell) for cell in row_tuple):
            data.append(tuple(row_tuple))
    return data


def create_data(list_1, list_2, keys):
    dictionary_list = []

    for index, amount in list_1:
        matched_tuple = list_2[index]

        dictionary = {llave: valor for llave, valor in zip(keys, matched_tuple)}
        dictionary["Cantidad"] = amount

        dictionary_list.append(dictionary)
    return dictionary_list


@app.route("/", methods=["GET", "POST"])
def front():
    data = []
    values = []
    totals = []
    elements = []
    food_table = []
    legends = [
        "energy",
        "protein",
        "fat",
        "calcium",
        "iron",
        "vitamin_a",
        "thiamine",
        "riboflavin",
        "niacin",
        "folate",
        "vitamin_c",
    ]
    labels = [
        "Alimento",
        "Categoria",
        "Energía (kcal)",
        "Proteína (g)",
        "Grasa (g)",
        "Calcio (mg)",
        "Hierro (mg)",
        "Vitamina A (µg)",
        "Tiamina (mg)",
        "Riboflavina (mg)",
        "Niacina (mg)",
        "Folato (µg)",
        "Vitamina C (mg)",
        "Cantidad",
    ]
    invalid = False
    delimiters = "[-,;_]"
    image_url = ""

    default_dataset_path = os.path.join(
        app.config["UPLOAD_FOLDER"], "dataset_alimentos.xlsx"
    )
    if not session.get("file_uploaded") and os.path.isfile(default_dataset_path):
        session["uploaded_file"] = default_dataset_path
        session["file_uploaded"] = True

    if request.method == "POST":
        if "file" in request.files:
            file = request.files["file"]
            if file and file.filename.endswith(".xlsx"):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                session["uploaded_file"] = filepath
                session["file_uploaded"] = True
                return "Correcto", 200
            else:
                return "Incorrecto", 400

        else:
            uploaded_filepath = session.get("uploaded_file")
            if uploaded_filepath and not os.path.isfile(uploaded_filepath):
                session.pop("uploaded_file", None)
                session["file_uploaded"] = False
                uploaded_filepath = None

            if uploaded_filepath is None:
                flash("Por favor, primero sube el dataset", "notify")
                return redirect(url_for("front"))

            food_table = process_file(uploaded_filepath)

            for i in range(1, 12):
                key = f"step_{i}"
                value = request.form.get(key, "0").strip()
                if value == "":
                    values.append(0)
                elif is_number(value):
                    values.append(float(value) if "." in value else int(value))
                else:
                    invalid = True
                    values.append(value)

            send_back = ["" if x == 0 else x for x in values]

            if invalid:
                flash("Por favor, ingresa solo números", "notify")
                flash(send_back, "text")
                return redirect(url_for("front"))

            all_zero = all(i == 0 for i in values)
            if all_zero:
                flash("Ingresa tus requerimientos", "notify")
                return redirect(url_for("front"))

            params = request.form.get("params", "50 , 100 , 0.8 , 0.05 , 5")
            raw_elements = re.split(delimiters, params)

            if len(raw_elements) != 5:
                flash("Ingresa 5 parámetros en orden", "notify")
                flash(send_back, "text")
                flash(params, "params")
                return redirect(url_for("front"))

            for i, v in enumerate(raw_elements):
                v = v.strip()

                if is_number(v):
                    if i < 2 or i == 4:
                        elements.append(int(float(v)))
                    else:
                        elements.append(float(v))
                else:
                    flash("Los parámetros deben ser números", "notify")
                    flash(send_back, "text")
                    flash(params, "params")
                    return redirect(url_for("front"))

            if elements[4] > elements[0]:
                flash("Torneo mayor que la población", "notify")
                flash(send_back, "text")
                flash(params, "params")
                return redirect(url_for("front"))

            (
                population_size,
                num_generations,
                crossover_prob,
                mutation_prob,
                tournament_size,
            ) = elements
            daily_requirements = dict(zip(legends, values))

            best_diet = genetic_algorithm.execute(
                food_table,
                daily_requirements,
                population_size,
                num_generations,
                crossover_prob,
                mutation_prob,
                tournament_size,
            )

            data = create_data(best_diet, food_table, labels)
            image_url = url_for("graph")

            totals = [0] * (len(list(data[0].values())) - 3)
            for d in data:
                vs = list(d.values())[2:-1]
                amount = list(d.values())[-1]
                for i, v in enumerate(vs):
                    totals[i] += v * amount
                    totals[i] = round(totals[i], 2)

            difference = [round(a - b, 2) for a, b in zip(totals, values)]

            flash(totals, "totals")
            flash(values, "required")
            flash(difference, "difference")
            flash(send_back, "text")
            flash(params, "params")
    return render_template(
        "front.html",
        file_uploaded=session.get("file_uploaded", False),
        data=data,
        imagen_url=image_url,
    )


@app.route("/grafica")
def graph():
    return send_file("graphs/fitness_plot.png", mimetype="image/png")


def main():
    app.run(debug=True, port="5000", host="0.0.0.0")


if __name__ == "__main__":
    main()
